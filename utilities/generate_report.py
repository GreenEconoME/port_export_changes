# Import dependencies
import time
import pandas as pd
import streamlit as st

from io import BytesIO
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Import helper functions
from utilities.differences import diff_dataframe
from utilities.formatting import highlight_changes, highlight_rows, display_progress_list



# Function to generate the report with progress display
def gen_report(unique_keys, sheets_to_compare, progress_placeholder, old, new):
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet in sheets_to_compare:
            if sheet not in old or sheet not in new:
                continue

            # Update sheet status to "Processing" (Orange)
            st.session_state["sheet_status"][sheet] = "Processing"
            progress_placeholder.markdown(display_progress_list(st.session_state["sheet_status"]))
            time.sleep(0.5)  # Delay to show UI update
            
            # Load sheet data
            old_df = old[sheet]
            new_df = new[sheet]
            
            diff_df = diff_dataframe(old_df, new_df, key_cols=unique_keys[sheet])
                    
            styled = diff_df.style \
                .apply(highlight_rows, axis=1) \
                .applymap(highlight_changes)
            
            styled.to_excel(writer, sheet_name=sheet, index=False)
                
            # Add a Table over that same shape
            row_count, col_count = diff_df.shape
                
            end_row = row_count + 1 
            end_col = col_count  
                
            start_cell = "A1"
            end_col_letter = get_column_letter(end_col)  # i.e. 'D' if col_count=4
            end_cell = f"{end_col_letter}{end_row}"
            table_ref = f"{start_cell}:{end_cell}"
                
            worksheet = writer.sheets[sheet]
                
            # Make a unique table name
            safe_sheet_name = sheet.replace(" ", "_").replace("-", "_")
            table_name = f"Table_{safe_sheet_name}"
                
            # Create the table
            tab = Table(displayName=table_name, ref=table_ref)
                
            # Use normal blue table style
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showRowStripes=True,
                showColumnStripes=False,
            )
            tab.tableStyleInfo = style
            worksheet.add_table(tab)
                
            # Fit column widths:
            for col_idx, col_cells in enumerate(worksheet.columns, start=1):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                    
                for cell in col_cells:
                    if cell.value:
                        cell_len = len(str(cell.value))
                        if cell_len > max_length:
                            max_length = cell_len
                    
                # Add padding for readability
                adjusted_width = max_length + 3
                worksheet.column_dimensions[col_letter].width = adjusted_width

            # Mark as Completed (Green)
            st.session_state["sheet_status"][sheet] = "Completed"
            progress_placeholder.markdown(display_progress_list(st.session_state["sheet_status"]))
            time.sleep(0.5)  # Delay to update UI

        # Set generating report to processing
        st.session_state["sheet_status"]["Generating Download"] = "Processing"
        progress_placeholder.markdown(display_progress_list(st.session_state["sheet_status"]))
        time.sleep(0.5)  # Delay to update UI
    # Move the buffer position to the beginning to allow reading it
    output.seek(0)
    
    return output